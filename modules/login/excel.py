import openpyxl
from openpyxl.styles import PatternFill
from modules.config.config import load_config
from modules.config import rules

config = load_config()
excel_file_path = config["excel_file_path"]


# 定义字段映射和处理函数
field_mapping = {
    "login_name": {
        "title": "学号",
        "function": rules.normal_rule,  # 使用 rules 模块中的规则
    },
    "password": {
        "title": "身份证",
        "function": rules.birthday_rule,  # 使用 rules 模块中的规则
    },
    # 可以额外添加更多的字段映射, 但需要同时修改rules.py
}


# 创建无色填充样式
def restore_red_cells_to_no_fill(sheet):
    no_fill = PatternFill(fill_type=None)
    red_fill = PatternFill(
        start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
    )
    for row in sheet.iter_rows():
        for cell in row:
            # 检查单元格是否为红色填充
            if cell.fill == red_fill:
                # 恢复为无色填充
                cell.fill = no_fill


def read_excel():
    users_account = []
    invalid_users = 0
    total_attempts = 0

    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        print(f"\n任务启动!\n成功加载工作簿: {excel_file_path}")

        # 创建红色填充样式
        red_fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )
        restore_red_cells_to_no_fill(sheet)
        workbook.save(excel_file_path)

        # 获取列标题, 保留所有列（包括None值）
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if not headers or all(header is None for header in headers):
            print(
                "(如果没有获取到任何结果,请将所有列标题放到第一行并且所有单元格取消合并)"
            )
            return None

        print(f"\n第一行获取到的列标题: {headers}")

        # 移除None值并检查是否有重复的列标题
        cleaned_headers = [header for header in headers if header is not None]
        header_counts = {
            header: cleaned_headers.count(header) for header in cleaned_headers
        }
        for header, count in header_counts.items():
            if count > 1:
                # 标记重复的列标题为红色
                for cell in sheet[1]:  # 第一行
                    if cell.value == header:
                        cell.fill = red_fill
                print(f"\n存在重复的列标题: '{header}', 请删除重复列后重试。")
                workbook.save(excel_file_path)
                return None

        # 确定列索引
        field_indices = {}
        for field, processor in field_mapping.items():
            try:
                field_indices[field] = headers.index(processor["title"])
            except ValueError:
                print(f"\n列标题中未找到字段: {processor['title']}")
                return None

        # 从表格第二行开始向下遍历数据
        for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            # 跳过完全为空的行
            if all(cell.value is None for cell in row):
                continue

            total_attempts += 1
            user_data = {}
            invalid_fields = []  # 用于记录无效的字段

            # 创建一个基于当前行数据的字典, 使用列标题作为键
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):  # 确保索引在列标题的范围内
                    row_data[headers[i]] = cell.value

            for field, processor in field_mapping.items():
                title = processor["title"]
                value = row_data.get(title)  # 获取当前行中对应列标题的数据
                # 确保传递给规则函数的值是字符串
                processed_value = processor["function"](
                    str(value) if value is not None else None
                )
                user_data[field] = processed_value

                # 如果处理后的值为None, 记录无效字段
                if processed_value is None:
                    # 找到当前行的单元格, 并记录下来
                    cell_index = headers.index(title)
                    if cell_index < len(row):  # 确保索引在当前行的范围内
                        invalid_fields.append(row[cell_index])

            # 如果用户名或密码为None, 则跳过当前行, 并设置单元格背景颜色为红色
            if user_data["login_name"] is None or user_data["password"] is None:
                invalid_users += 1
                for cell in invalid_fields:
                    cell.fill = red_fill
                continue

            users_account.append(user_data)

        # 保存工作簿
        workbook.save(excel_file_path)

    except FileNotFoundError:
        print(f"\n文件未找到: {excel_file_path}")
    except Exception as e:
        print(f"\n加载Excel文件时发生错误: {e}")
        return None

    # 输出获取成功的用户数量, 无效的用户数量, 尝试获取的总数
    print(
        f"\n√√√ 用户读取成功数量: [{len(users_account)}/{total_attempts}]"
        f"\nxxx 错误数量: [{invalid_users}/{total_attempts}]"
    )

    return users_account
